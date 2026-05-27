import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import ttkbootstrap as ttkb
import pandas as pd
import sqlite3
import io
import os
import sys
from datetime import datetime


# ── Configuration ──────────────────────────────────────────────────────────────
NETWORK_PATH = r"\\sidel.com\emea\pt-smf\groups\STORAGE\Máquinas\Produção\caps"
DB_PATH      = os.path.join(NETWORK_PATH, 'capsulas.db')

ESTADO_OPTIONS = ['Novo', 'Usado', 'Amostras', 'Stock', 'S/ID', 'NOK', 'Outros']

ESTADO_COLORS = {
    'Novo':     '#d4edda',
    'Usado':    '#fff3cd',
    'Amostras': '#cce5ff',
    'NOK':      '#f8d7da',
    'Stock':    '#e2e3e5',
    'S/ID':     '#fce8b2',
    'Outros':   '#e8d5f5',
}

COL_DISPLAY = ['ID', 'Localização', 'Projeto', 'Código', 'Data receção', 'Qtd.', 'Estado', 'Obs.']
COL_KEYS    = ['id', 'localizacao', 'projeto', 'codigo', 'data_rececao', 'qtd', 'estado', 'obs']
COL_WIDTHS  = [45, 120, 200, 130, 100, 55, 85, 0]  # 0 = stretch

SORT_MAP = {
    'ID': 'id', 'Localização': 'localizacao', 'Projeto': 'projeto',
    'Estado': 'estado', 'Código': 'codigo', 'Data receção': 'data_rececao', 'Qtd.': 'qtd',
}


# ── Database ───────────────────────────────────────────────────────────────────
def get_conn():
    return sqlite3.connect(DB_PATH, check_same_thread=False, timeout=10)


def init_db():
    with get_conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS capsulas (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                localizacao  TEXT,
                projeto      TEXT,
                codigo       TEXT,
                data_rececao TEXT,
                qtd          TEXT,
                estado       TEXT,
                obs          TEXT,
                source       TEXT DEFAULT 'excel',
                created_at   TEXT DEFAULT (datetime('now')),
                updated_at   TEXT DEFAULT (datetime('now'))
            )
        """)
        try:
            conn.execute("ALTER TABLE capsulas ADD COLUMN source TEXT DEFAULT 'excel'")
        except Exception:
            pass
        conn.commit()


def db_is_empty():
    with get_conn() as conn:
        return conn.execute("SELECT COUNT(*) FROM capsulas").fetchone()[0] == 0


def load_all():
    with get_conn() as conn:
        return pd.read_sql("SELECT * FROM capsulas ORDER BY id", conn)


def insert_row(d):
    with get_conn() as conn:
        conn.execute(
            "INSERT INTO capsulas (localizacao,projeto,codigo,data_rececao,qtd,estado,obs,source) "
            "VALUES (?,?,?,?,?,?,?,'manual')",
            (d['localizacao'], d['projeto'], d['codigo'],
             d['data_rececao'], d['qtd'], d['estado'], d['obs']),
        )
        conn.commit()


def update_row(row_id, d):
    with get_conn() as conn:
        conn.execute(
            "UPDATE capsulas SET localizacao=?,projeto=?,codigo=?,data_rececao=?,qtd=?,"
            "estado=?,obs=?,updated_at=datetime('now') WHERE id=?",
            (d['localizacao'], d['projeto'], d['codigo'],
             d['data_rececao'], d['qtd'], d['estado'], d['obs'], row_id),
        )
        conn.commit()


def delete_row(row_id):
    with get_conn() as conn:
        conn.execute("DELETE FROM capsulas WHERE id=?", (row_id,))
        conn.commit()


def bulk_import(df):
    with get_conn() as conn:
        conn.execute("DELETE FROM capsulas WHERE source='excel' OR source IS NULL")
        max_manual = conn.execute(
            "SELECT COALESCE(MAX(id), 0) FROM capsulas WHERE source='manual'"
        ).fetchone()[0]
        conn.execute("DELETE FROM sqlite_sequence WHERE name='capsulas'")
        conn.execute("INSERT INTO sqlite_sequence (name, seq) VALUES ('capsulas', ?)", (max_manual,))
        for _, r in df.iterrows():
            conn.execute(
                "INSERT INTO capsulas (localizacao,projeto,codigo,data_rececao,qtd,estado,obs,source) "
                "VALUES (?,?,?,?,?,?,?,'excel')",
                (_safe(r, 'localizacao'), _safe(r, 'projeto'), _safe(r, 'codigo'),
                 _safe(r, 'data_rececao'), _safe(r, 'qtd'), _safe(r, 'estado'), _safe(r, 'obs')),
            )
        conn.commit()


def _safe(row, col):
    v = row.get(col, '')
    return '' if pd.isna(v) or str(v) in ('nan', 'NaT', 'None') else str(v).strip()


# ── Excel helpers ──────────────────────────────────────────────────────────────
def _normalize_estado(v):
    if pd.isna(v) or str(v).strip() in ('', 'nan'):
        return ''
    mapping = {
        'novo': 'Novo', 'novas': 'Novo', 'new': 'Novo',
        'usado': 'Usado', 'usadas': 'Usado', 'usadas e novas': 'Usado',
        'amostras': 'Amostras', 'amostra': 'Amostras',
        's/id': 'S/ID', 'sem id': 'S/ID', 'sem identificação': 'S/ID',
        'stock': 'Stock', 'nok': 'NOK', 'outros': 'Outros',
    }
    return mapping.get(str(v).lower().strip(), str(v).strip())


def _detect_col(columns):
    targets = {
        'localizacao':  ['local', 'loca'],
        'projeto':      ['projeto', 'project'],
        'codigo':       ['digo', 'cod', 'code'],
        'data_rececao': ['data', 'rece'],
        'qtd':          ['qtd', 'quant'],
        'estado':       ['estado', 'state', 'status'],
        'obs':          ['obs'],
    }
    result = {}
    for col in columns:
        col_l = col.lower()
        for internal, patterns in targets.items():
            if internal not in result and any(p in col_l for p in patterns):
                result[col] = internal
                break
    return result


def read_excel(file):
    xl = pd.ExcelFile(file, engine='openpyxl')
    df = xl.parse(xl.sheet_names[0])
    df = df.rename(columns=_detect_col(df.columns.tolist()))
    if 'estado' in df.columns:
        df['estado'] = df['estado'].apply(_normalize_estado)
    for col in ['localizacao', 'projeto', 'codigo', 'data_rececao', 'qtd', 'estado', 'obs']:
        if col not in df.columns:
            df[col] = ''
    return df[['localizacao', 'projeto', 'codigo', 'data_rececao', 'qtd', 'estado', 'obs']]


def find_excel_in_folder():
    try:
        for f in os.listdir(NETWORK_PATH):
            if f.lower().endswith('.xlsx'):
                return os.path.join(NETWORK_PATH, f)
    except OSError:
        pass
    return None


def to_excel_bytes(df):
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    out = io.BytesIO()
    export_df = df[['localizacao', 'projeto', 'codigo', 'data_rececao', 'qtd', 'estado', 'obs']].copy()
    export_df.columns = ['Localização', 'Projeto', 'Código', 'Data receção', 'Qtd.', 'Estado', 'Obs.']
    export_df = export_df.replace('', pd.NA)

    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        export_df.to_excel(writer, sheet_name='Cápsulas', index=False)
        ws = writer.sheets['Cápsulas']

        hdr_fill = PatternFill('solid', fgColor='1F3864')
        hdr_font = Font(color='FFFFFF', bold=True, size=11)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        color_map = {k: v.lstrip('#') for k, v in ESTADO_COLORS.items()}
        for row_idx in range(2, ws.max_row + 1):
            estado_val = str(ws.cell(row=row_idx, column=6).value or '')
            if estado_val in color_map:
                fill = PatternFill('solid', fgColor=color_map[estado_val])
                for col_idx in range(1, 8):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

        for i, w in enumerate([12, 28, 16, 14, 6, 12, 45], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = 'A2'

    return out.getvalue()


# ── Form Dialog ────────────────────────────────────────────────────────────────
class FormDialog(tk.Toplevel):
    def __init__(self, parent, title, defaults=None, on_save=None):
        super().__init__(parent)
        self.title(title)
        self.on_save = on_save
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        d = defaults or {}
        p = {'padx': 8, 'pady': 5}

        frm = ttk.Frame(self, padding=16)
        frm.pack(fill=tk.BOTH, expand=True)

        # Row 0 — Localização + Projeto
        ttk.Label(frm, text='Localização *').grid(row=0, column=0, sticky='w', **p)
        self.v_loc = tk.StringVar(value=d.get('localizacao', ''))
        ttk.Entry(frm, textvariable=self.v_loc, width=22).grid(row=0, column=1, sticky='ew', **p)

        ttk.Label(frm, text='Projeto').grid(row=0, column=2, sticky='w', **p)
        self.v_proj = tk.StringVar(value=d.get('projeto', ''))
        ttk.Entry(frm, textvariable=self.v_proj, width=32).grid(row=0, column=3, columnspan=3, sticky='ew', **p)

        # Row 1 — Código + Data + Qtd + Estado
        ttk.Label(frm, text='Código').grid(row=1, column=0, sticky='w', **p)
        self.v_cod = tk.StringVar(value=d.get('codigo', ''))
        ttk.Entry(frm, textvariable=self.v_cod, width=22).grid(row=1, column=1, sticky='ew', **p)

        ttk.Label(frm, text='Data receção').grid(row=1, column=2, sticky='w', **p)
        self.v_data = tk.StringVar(value=d.get('data_rececao', ''))
        ttk.Entry(frm, textvariable=self.v_data, width=14).grid(row=1, column=3, sticky='ew', **p)

        ttk.Label(frm, text='Qtd.').grid(row=1, column=4, sticky='w', **p)
        self.v_qtd = tk.StringVar(value=d.get('qtd', ''))
        ttk.Entry(frm, textvariable=self.v_qtd, width=8).grid(row=1, column=5, sticky='ew', **p)

        # Row 2 — Estado
        ttk.Label(frm, text='Estado').grid(row=2, column=0, sticky='w', **p)
        self.v_estado = tk.StringVar(value=d.get('estado', ESTADO_OPTIONS[0]))
        cb = ttk.Combobox(frm, textvariable=self.v_estado, values=ESTADO_OPTIONS,
                          state='readonly', width=14)
        cb.grid(row=2, column=1, sticky='w', **p)

        # Row 3 — Obs
        ttk.Label(frm, text='Observações').grid(row=3, column=0, sticky='nw', **p)
        self.txt_obs = tk.Text(frm, height=3, width=62, font=('Segoe UI', 9))
        self.txt_obs.grid(row=3, column=1, columnspan=5, sticky='ew', **p)
        self.txt_obs.insert('1.0', d.get('obs', ''))

        # Buttons
        btn_frm = ttk.Frame(frm)
        btn_frm.grid(row=4, column=0, columnspan=6, sticky='w', pady=(10, 0))
        ttk.Button(btn_frm, text='💾  Guardar', command=self._save,
                   bootstyle='success').pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(btn_frm, text='Cancelar', command=self.destroy,
                   bootstyle='secondary').pack(side=tk.LEFT)

        self.update_idletasks()
        pw, ph = parent.winfo_width(), parent.winfo_height()
        px, py = parent.winfo_x(), parent.winfo_y()
        dw, dh = self.winfo_width(), self.winfo_height()
        self.geometry(f'+{px + (pw - dw) // 2}+{py + (ph - dh) // 2}')

        self.bind('<Return>', lambda e: self._save())
        self.bind('<Escape>', lambda e: self.destroy())
        self.wait_window()

    def _save(self):
        loc = self.v_loc.get().strip()
        if not loc:
            messagebox.showerror('Erro', 'A Localização é obrigatória.', parent=self)
            return
        d = {
            'localizacao': loc,
            'projeto':     self.v_proj.get().strip(),
            'codigo':      self.v_cod.get().strip(),
            'data_rececao': self.v_data.get().strip(),
            'qtd':         self.v_qtd.get().strip(),
            'estado':      self.v_estado.get(),
            'obs':         self.txt_obs.get('1.0', tk.END).strip(),
        }
        if self.on_save:
            self.on_save(d)
        self.destroy()


# ── Main Application ───────────────────────────────────────────────────────────
class App(ttkb.Window):
    def __init__(self):
        super().__init__(
            title='Gestão de Stock de Cápsulas',
            themename='cosmo',
            minsize=(1100, 640),
        )
        self._df = pd.DataFrame()
        self._build_ui()
        self._refresh()
        self.state('zoomed')

    def _build_ui(self):
        # ── Header ──────────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg='#1F3864', pady=10)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text='  Gestão de Stock de Cápsulas',
                 bg='#1F3864', fg='white',
                 font=('Segoe UI', 15, 'bold')).pack(side=tk.LEFT, padx=8)
        tk.Label(hdr, text='Visualize, edite e exporte o inventário de cápsulas',
                 bg='#1F3864', fg='#aaccff',
                 font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=4)

        # ── Toolbar ─────────────────────────────────────────────────────────────
        tb = ttk.Frame(self, padding=(10, 6))
        tb.pack(fill=tk.X)
        ttk.Button(tb, text='🔄  Reload Excel',   command=self._reload_excel,
                   bootstyle='primary', width=17).pack(side=tk.LEFT, padx=3)
        ttk.Button(tb, text='➕  Nova Cápsula',   command=self._add_row,
                   bootstyle='success', width=17).pack(side=tk.LEFT, padx=3)
        ttk.Button(tb, text='📤  Exportar Excel', command=self._export_excel,
                   bootstyle='info',    width=17).pack(side=tk.LEFT, padx=3)

        # ── Filters ─────────────────────────────────────────────────────────────
        flt = ttk.LabelFrame(self, text='Filtros', padding=(10, 5))
        flt.pack(fill=tk.X, padx=10, pady=(2, 0))

        ttk.Label(flt, text='Pesquisar:').pack(side=tk.LEFT, padx=(0, 3))
        self.v_search = tk.StringVar()
        self.v_search.trace_add('write', lambda *_: self._apply_filters())
        ttk.Entry(flt, textvariable=self.v_search, width=26).pack(side=tk.LEFT, padx=(0, 14))

        ttk.Label(flt, text='Estado:').pack(side=tk.LEFT, padx=(0, 3))
        self.v_estado_f = tk.StringVar(value='Todos')
        self.v_estado_f.trace_add('write', lambda *_: self._apply_filters())
        ttk.Combobox(flt, textvariable=self.v_estado_f,
                     values=['Todos'] + ESTADO_OPTIONS,
                     state='readonly', width=12).pack(side=tk.LEFT, padx=(0, 14))

        ttk.Label(flt, text='Ordenar por:').pack(side=tk.LEFT, padx=(0, 3))
        self.v_sort = tk.StringVar(value='Localização')
        self.v_sort.trace_add('write', lambda *_: self._apply_filters())
        ttk.Combobox(flt, textvariable=self.v_sort,
                     values=list(SORT_MAP.keys()),
                     state='readonly', width=14).pack(side=tk.LEFT, padx=(0, 6))

        self.v_asc = tk.BooleanVar(value=True)
        ttk.Checkbutton(flt, text='Crescente', variable=self.v_asc,
                        command=self._apply_filters,
                        bootstyle='round-toggle').pack(side=tk.LEFT, padx=4)

        # ── Stat cards ──────────────────────────────────────────────────────────
        sf = ttk.Frame(self, padding=(10, 6))
        sf.pack(fill=tk.X)
        self._stat_vals = {}
        for key in ['Registos', 'Total cápsulas', 'Localizações', 'Projetos', 'Total na BD']:
            card = tk.Frame(sf, bg='#1558b0', padx=14, pady=8)
            card.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
            vl = tk.Label(card, text='—', font=('Segoe UI', 20, 'bold'),
                          bg='#1558b0', fg='white')
            vl.pack()
            tk.Label(card, text=key, font=('Segoe UI', 8),
                     bg='#1558b0', fg='#cce0ff').pack()
            self._stat_vals[key] = vl

        # ── Table ───────────────────────────────────────────────────────────────
        tbl = ttk.Frame(self, padding=(10, 2))
        tbl.pack(fill=tk.BOTH, expand=True)

        style = ttk.Style()
        style.configure('Treeview', rowheight=22, font=('Segoe UI', 9))
        style.configure('Treeview.Heading', font=('Segoe UI', 9, 'bold'))

        self.tree = ttk.Treeview(tbl, columns=COL_DISPLAY, show='headings',
                                  selectmode='browse')

        for col, width in zip(COL_DISPLAY, COL_WIDTHS):
            anchor = 'center' if col in ('ID', 'Qtd.') else 'w'
            stretch = col == 'Obs.'
            self.tree.heading(col, text=col,
                              command=lambda c=col: self._col_sort(c))
            self.tree.column(col, width=width, anchor=anchor,
                             minwidth=30, stretch=stretch)

        for estado, color in ESTADO_COLORS.items():
            self.tree.tag_configure(estado, background=color)

        vsb = ttk.Scrollbar(tbl, orient='vertical',   command=self.tree.yview)
        hsb = ttk.Scrollbar(tbl, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side=tk.RIGHT,  fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.pack(fill=tk.BOTH, expand=True)

        self.tree.bind('<<TreeviewSelect>>', self._on_select)
        self.tree.bind('<Double-1>',         self._on_double_click)

        # Context menu
        self._ctx = tk.Menu(self, tearoff=0)
        self._ctx.add_command(label='✏️  Editar',  command=self._edit_selected)
        self._ctx.add_command(label='🗑️  Apagar', command=self._delete_selected)
        self.tree.bind('<Button-3>', self._show_ctx)

        # ── Bottom bar ──────────────────────────────────────────────────────────
        bot = ttk.Frame(self, padding=(10, 5))
        bot.pack(fill=tk.X)

        ttk.Label(bot, text='ID do registo:').pack(side=tk.LEFT, padx=(0, 4))
        self.v_id = tk.IntVar(value=1)
        ttk.Spinbox(bot, from_=1, to=999999, textvariable=self.v_id,
                    width=7).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(bot, text='✏️  Editar', command=self._edit_by_id,
                   bootstyle='warning', width=12).pack(side=tk.LEFT, padx=3)
        ttk.Button(bot, text='🗑️  Apagar', command=self._delete_by_id,
                   bootstyle='danger',  width=12).pack(side=tk.LEFT, padx=3)

        self.v_status = tk.StringVar()
        ttk.Label(bot, textvariable=self.v_status,
                  font=('Segoe UI', 9)).pack(side=tk.RIGHT, padx=10)

    # ── Data ────────────────────────────────────────────────────────────────────
    def _refresh(self):
        self._df = load_all()
        if db_is_empty():
            excel_path = find_excel_in_folder()
            if excel_path:
                try:
                    bulk_import(read_excel(excel_path))
                    self._df = load_all()
                except Exception:
                    pass
        self._apply_filters()

    def _apply_filters(self):
        df = self._df.copy() if not self._df.empty else pd.DataFrame(columns=COL_KEYS)

        q = self.v_search.get().strip().lower()
        if q and not df.empty:
            mask = (
                df['localizacao'].str.lower().str.contains(q, na=False) |
                df['projeto'].str.lower().str.contains(q, na=False)     |
                df['codigo'].str.lower().str.contains(q, na=False)      |
                df['obs'].str.lower().str.contains(q, na=False)
            )
            df = df[mask]

        estado_f = self.v_estado_f.get()
        if estado_f != 'Todos' and not df.empty:
            df = df[df['estado'] == estado_f]

        sort_col = SORT_MAP.get(self.v_sort.get(), 'localizacao')
        if sort_col in df.columns and not df.empty:
            df = df.sort_values(sort_col, ascending=self.v_asc.get(), na_position='last')

        self._populate_table(df)
        self._update_stats(df)

    def _populate_table(self, df):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for _, row in df.iterrows():
            vals = ['' if str(row.get(k, '')) in ('nan', 'NaT', 'None', 'nan')
                    else row.get(k, '') for k in COL_KEYS]
            estado = str(row.get('estado', ''))
            tag = (estado,) if estado in ESTADO_COLORS else ()
            self.tree.insert('', tk.END, values=vals, tags=tag)

    def _update_stats(self, df):
        total_qtd = pd.to_numeric(df['qtd'], errors='coerce').sum() if not df.empty else 0
        self._stat_vals['Registos'].config(text=str(len(df)))
        self._stat_vals['Total cápsulas'].config(
            text=str(int(total_qtd)) if pd.notna(total_qtd) and total_qtd > 0 else '—')
        self._stat_vals['Localizações'].config(
            text=str(df['localizacao'].nunique()) if not df.empty else '0')
        self._stat_vals['Projetos'].config(
            text=str(df['projeto'].nunique()) if not df.empty else '0')
        self._stat_vals['Total na BD'].config(text=str(len(self._df)))

    # ── Actions ─────────────────────────────────────────────────────────────────
    def _reload_excel(self):
        excel_path = find_excel_in_folder()
        if not excel_path:
            excel_path = filedialog.askopenfilename(
                title='Selecionar ficheiro Excel',
                filetypes=[('Excel', '*.xlsx *.xls')],
                initialdir=NETWORK_PATH,
            )
        if not excel_path:
            return
        try:
            bulk_import(read_excel(excel_path))
            self._df = load_all()
            self._apply_filters()
            self._set_status(f'✅ Excel recarregado — {len(self._df)} registos')
        except Exception as e:
            messagebox.showerror('Erro ao importar', str(e))

    def _add_row(self):
        def on_save(d):
            insert_row(d)
            self._df = load_all()
            self._apply_filters()
            self._set_status('✅ Cápsula adicionada.')
        FormDialog(self, 'Adicionar nova cápsula', on_save=on_save)

    def _edit_row(self, row_id):
        all_data = load_all()
        match = all_data[all_data['id'] == row_id]
        if match.empty:
            messagebox.showerror('Erro', f'ID {row_id} não existe.')
            return
        def on_save(d):
            update_row(row_id, d)
            self._df = load_all()
            self._apply_filters()
            self._set_status(f'✅ Registo {row_id} atualizado.')
        FormDialog(self, f'Editar registo ID {row_id}',
                   defaults=match.iloc[0].to_dict(), on_save=on_save)

    def _delete_row(self, row_id):
        all_data = load_all()
        match = all_data[all_data['id'] == row_id]
        if match.empty:
            messagebox.showerror('Erro', f'ID {row_id} não existe.')
            return
        r = match.iloc[0]
        if messagebox.askyesno('Confirmar eliminação',
                               f'Eliminar registo ID {row_id}?\n\n'
                               f'{r["localizacao"]} / {r["projeto"]}'):
            delete_row(row_id)
            self._df = load_all()
            self._apply_filters()
            self._set_status(f'Registo {row_id} eliminado.')

    def _edit_by_id(self):
        self._edit_row(self.v_id.get())

    def _delete_by_id(self):
        self._delete_row(self.v_id.get())

    def _edit_selected(self):
        item = self.tree.focus()
        if item:
            self._edit_row(int(self.tree.item(item)['values'][0]))

    def _delete_selected(self):
        item = self.tree.focus()
        if item:
            self._delete_row(int(self.tree.item(item)['values'][0]))

    def _export_excel(self):
        if self._df.empty:
            messagebox.showinfo('Info', 'Sem dados para exportar.')
            return
        filename = f'capsulas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            initialfile=filename,
            filetypes=[('Excel', '*.xlsx')],
            initialdir=NETWORK_PATH,
        )
        if path:
            with open(path, 'wb') as f:
                f.write(to_excel_bytes(self._df))
            self._set_status(f'✅ Exportado: {os.path.basename(path)}')

    # ── Helpers ─────────────────────────────────────────────────────────────────
    def _on_select(self, _event):
        item = self.tree.focus()
        if item:
            self.v_id.set(int(self.tree.item(item)['values'][0]))

    def _on_double_click(self, _event):
        self._edit_selected()

    def _show_ctx(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self.tree.focus(item)
            self._ctx.post(event.x_root, event.y_root)

    def _col_sort(self, col):
        if self.v_sort.get() == col:
            self.v_asc.set(not self.v_asc.get())
        else:
            self.v_sort.set(col)
            self.v_asc.set(True)
        self._apply_filters()

    def _set_status(self, msg):
        self.v_status.set(msg)
        self.after(4000, lambda: self.v_status.set(''))


# ── Entry point ────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    if not os.path.isdir(NETWORK_PATH):
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror(
            'Rede não disponível',
            f'Não foi possível aceder à pasta partilhada:\n\n'
            f'{NETWORK_PATH}\n\n'
            'Verifique se está ligado à rede da empresa e tente novamente.'
        )
        root.destroy()
        sys.exit(1)

    init_db()
    App().mainloop()
